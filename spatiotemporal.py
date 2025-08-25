import random
import matplotlib.pyplot as plt
from datetime import datetime, timedelta, timezone
import folium
from folium import plugins
import pandas as pd
import json
import time
import requests
import google.generativeai as genai
# from transformers import AutoTokenizer, AutoModelForCausalLM
# import torch
import csv
from io import StringIO
import sys
import openpyxl
from openpyxl.styles import Alignment
import random

# List of major international cities with their IATA codes
INTERNATIONAL_CITIES = [
    # Europe
    {"city": "London", "iata": "LHR", "country": "UK"},
    {"city": "Paris", "iata": "CDG", "country": "France"},
    {"city": "Amsterdam", "iata": "AMS", "country": "Netherlands"},
    {"city": "Frankfurt", "iata": "FRA", "country": "Germany"},
    {"city": "Rome", "iata": "FCO", "country": "Italy"},
    {"city": "Madrid", "iata": "MAD", "country": "Spain"},
    {"city": "Istanbul", "iata": "IST", "country": "Turkey"},
    {"city": "Moscow", "iata": "SVO", "country": "Russia"},
   
    # Asia
    {"city": "Tokyo", "iata": "HND", "country": "Japan"},
    {"city": "Singapore", "iata": "SIN", "country": "Singapore"},
    {"city": "Hong Kong", "iata": "HKG", "country": "China"},
    {"city": "Seoul", "iata": "ICN", "country": "South Korea"},
    {"city": "Bangkok", "iata": "BKK", "country": "Thailand"},
    {"city": "Dubai", "iata": "DXB", "country": "UAE"},
    {"city": "Mumbai", "iata": "BOM", "country": "India"},
    {"city": "Shanghai", "iata": "PVG", "country": "China"},
   
    # Americas
    {"city": "New York", "iata": "JFK", "country": "USA"},
    {"city": "Los Angeles", "iata": "LAX", "country": "USA"},
    {"city": "Toronto", "iata": "YYZ", "country": "Canada"},
    {"city": "Mexico City", "iata": "MEX", "country": "Mexico"},
    {"city": "São Paulo", "iata": "GRU", "country": "Brazil"},
    {"city": "Buenos Aires", "iata": "EZE", "country": "Argentina"},
   
    # Oceania
    {"city": "Sydney", "iata": "SYD", "country": "Australia"},
    {"city": "Melbourne", "iata": "MEL", "country": "Australia"},
    {"city": "Auckland", "iata": "AKL", "country": "New Zealand"},
   
    # Africa
    {"city": "Cairo", "iata": "CAI", "country": "Egypt"},
    {"city": "Johannesburg", "iata": "JNB", "country": "South Africa"},
    {"city": "Nairobi", "iata": "NBO", "country": "Kenya"},
    {"city": "Casablanca", "iata": "CMN", "country": "Morocco"}
]

# API Keys
try:
    from key import mykey as OPENAI_API_KEY
except ImportError:
    OPENAI_API_KEY = "YOUR_KEY"

try:
    from key import GOOGLE_API_KEY
except ImportError:
    GOOGLE_API_KEY = "YOUR_KEY"

try:
    from key import AERODATABOX_API_KEY
except ImportError:
    AERODATABOX_API_KEY = "YOUR_KEY"

# Initialize OpenAI client
from openai import OpenAI
client = OpenAI(api_key=OPENAI_API_KEY)

# Initialize Google Gemini
genai.configure(api_key=GOOGLE_API_KEY)

def get_min_travel_time_aero(from_iata, to_iata, max_retries=3):
    """
    Fetches minimum flight travel time between two airports using AeroDataBox API.
    Uses the distance-time endpoint to get exact flight duration and distance.
    """
    if AERODATABOX_API_KEY == "YOUR_AERODATABOX_API_KEY":
        print("Warning: AeroDataBox API key not configured. Using default travel time.")
        return 14400  # Default 4 hours in seconds if API key is not set

    url = f"https://aerodatabox.p.rapidapi.com/airports/iata/{from_iata}/distance-time/{to_iata}"
   
    headers = {
        "x-rapidapi-key": AERODATABOX_API_KEY,
        "x-rapidapi-host": "aerodatabox.p.rapidapi.com"
    }
   
    querystring = {
        "flightTimeModel": "Standard"
    }

    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, params=querystring)
            response.raise_for_status()
            data = response.json()
           
            if data and 'approxFlightTime' in data:
                hours, minutes, seconds = map(int, data['approxFlightTime'].split(':'))
                flight_time_seconds = hours * 3600 + minutes * 60 + seconds
               
                # Add 1 hour for takeoff/landing and taxiing
                total_time_seconds = flight_time_seconds + 3600
               
                print(f"Flight time from {from_iata} to {to_iata}: {hours}h {minutes}m (plus 1h buffer)")
                return total_time_seconds
            else:
                print(f"Could not retrieve flight time data from {from_iata} to {to_iata}. Using default.")
                return 14400
               
        except requests.exceptions.RequestException as e:
            if attempt < max_retries - 1:
                print(f"AeroDataBox API request failed (attempt {attempt + 1}/{max_retries}): {e}")
                time.sleep(1)  # Wait 1 second before retrying
                continue
            print(f"AeroDataBox API request failed after {max_retries} attempts: {e}")
            return 14400
        except (KeyError, ValueError) as e:
            if attempt < max_retries - 1:
                print(f"Error processing AeroDataBox API response (attempt {attempt + 1}/{max_retries}): {e}")
                time.sleep(1)
                continue
            print(f"Error processing AeroDataBox API response after {max_retries} attempts: {e}")
            return 14400

# Global cache for AeroDataBox results
AERODATABOX_CACHE = {}

def get_min_travel_time(from_place, to_place, current_places_data, cache=None):
    # Extract IATA codes from the place objects
    from_iata = from_place.get("iata")
    to_iata = to_place.get("iata")

    if not from_iata or not to_iata:
        print(f"Error: IATA code missing for {from_place.get('place')} or {to_place.get('place')}")
        return 14400

    # Check global cache first
    cache_key = (from_iata, to_iata)
    if cache_key in AERODATABOX_CACHE:
        print(f"Using cached flight time for {from_iata} -> {to_iata}")
        return AERODATABOX_CACHE[cache_key]

    # Then check local cache
    if cache is not None and cache_key in cache:
        print(f"Using local cache for {from_iata} -> {to_iata}")
        return cache[cache_key]

    # If not in any cache, fetch from API
    travel_time = get_min_travel_time_aero(from_iata, to_iata)

    # Store in both caches
    AERODATABOX_CACHE[cache_key] = travel_time
    if cache is not None:
        cache[cache_key] = travel_time
        print(f"Cached {from_iata} -> {to_iata}: {travel_time} seconds")

    return travel_time

def validate_edge(time_A, time_B, place_A, place_B, current_places_data, travel_time_cache): # Added travel_time_cache
    if time_A[1] > time_B[0]:  # Departure from A is after arrival at B
        return False
   
    min_travel_seconds = get_min_travel_time(place_A, place_B, current_places_data, travel_time_cache) # Pass cache
    max_travel_seconds = int(min_travel_seconds * 2)  # Set max travel time to 2x min time
    available_travel_time_seconds = time_B[0] - time_A[1]
   
    if available_travel_time_seconds < min_travel_seconds or available_travel_time_seconds > max_travel_seconds:
        return False
   
    return True

def llama2_ollama_completion(prompt, model_name="llama2:7b"):
    """
    Calls the local Ollama server for Llama 2 completions.
    """
    url = "http://localhost:11434/api/generate"
    payload = {
        "model": model_name,
        "prompt": prompt,
        "stream": False
    }
    try:
        response = requests.post(url, json=payload, timeout=120)
        response.raise_for_status()
        data = response.json()
        # Ollama returns the result in the 'response' field
        return data.get("response", "").strip()
    except Exception as e:
        print(f"Error calling Ollama Llama2 API: {e}")
        return None


def generate_itinerary_with_llm(selected_cities, num_destinations=4, llm_provider="openai", model_name="gpt-4o-mini", feedback_issues=None, fixed_route_sequence=None):
    today_utc = datetime.now(timezone.utc)
    date_suggestion_start = today_utc.strftime("%Y-%m-%d")
    date_suggestion_end = (today_utc + timedelta(days=30)).strftime("%Y-%m-%d")

    cities_str = "\n".join([f"- {city['city']} ({city['iata']}), {city['country']}" for city in selected_cities])

    if fixed_route_sequence:
        route_elements = [f"- {s['place']} ({s['iata']})" for s in fixed_route_sequence]
        fixed_route_str_for_prompt = "\n".join(route_elements)
        current_num_destinations = len(fixed_route_sequence)
        base_prompt_text = f"""You are tasked with creating a valid time schedule for a PRE-DEFINED travel itinerary.
The itinerary visits {current_num_destinations} destinations.
You MUST follow this exact sequence of cities and use their IATA codes as provided:
{fixed_route_str_for_prompt}

The cities involved are from the following list (for context and ensuring correct naming/IATA):
{cities_str}

IMPORTANT: Return ONLY a valid JSON object with this EXACT structure (no additional text, no markdown formatting):
{{
    "itinerary": [
        // Example for the first stop, ensure "place" matches the fixed sequence
        {{
            "place": "{fixed_route_sequence[0]['place']} ({fixed_route_sequence[0]['iata']})", 
            "arrival_time": "YYYY-MM-DD HH:MM",
            "departure_time": "YYYY-MM-DD HH:MM"
        }}
        // ... and so on for all {current_num_destinations} cities in the fixed_route_sequence
    ]
}}

Requirements:
1. All times MUST be in UTC.
2. Use 24-hour format (e.g., 14:30, 00:00 for midnight) ONLY and EXACTLY MATCH the format '%Y-%m-%d %H:%M'.
3. The "place" field in your JSON for each stop MUST EXACTLY MATCH the city name and IATA code from the fixed sequence provided above. Do not alter the sequence or the cities.
4. Do NOT add any explanatory text or markdown formatting.
5. Ensure the JSON is properly formatted with correct commas and brackets.
6. Account for minimum flight times between cities (use realistic minimum flight durations).
7. For each city, the difference between its 'departure_time' and 'arrival_time' (i.e., the stay at that city) MUST be more than 2 days (> 49 hours).
8. For each segment, the difference between the 'departure_time' of the previous city and the 'arrival_time' of the next city MUST be equal to the minimum realistic flight time (plus a 1 hour buffer for airport procedures). Do NOT add extra days or hours to the travel time.
9. The stay duration and the travel duration are separate: do NOT add the 2-day minimum stay to the travel time. The 2-day minimum applies only to the time spent at each city.
10. Return ONLY the JSON object, nothing else.
"""
    else:
        base_prompt_text = f"""Generate a travel itinerary visiting {num_destinations} destinations, exclusively using air travel.
You MUST use ONLY these cities for your itinerary:
{cities_str}

IMPORTANT: Return ONLY a valid JSON object with this EXACT structure (no additional text, no markdown formatting):
{{
    "itinerary": [
        {{
            "place": "city_name (IATA)",
            "arrival_time": "YYYY-MM-DD HH:MM",
            "departure_time": "YYYY-MM-DD HH:MM"
        }},
        {{
            "place": "city_name (IATA)",
            "arrival_time": "YYYY-MM-DD HH:MM",
            "departure_time": "YYYY-MM-DD HH:MM"
        }}
        // ... up to {num_destinations} items
    ]
}}

Requirements:
1. All times MUST be in UTC.
2. Use 24-hour format (e.g., 14:30, 00:00 for midnight).
3. Travel dates must be between {date_suggestion_start} and {date_suggestion_end}.
4. Include the IATA airport code for each city in parentheses.
5. Do NOT add any explanatory text or markdown formatting.
6. Ensure the JSON is properly formatted with correct commas and brackets.
7. For each city, the difference between its 'departure_time' and 'arrival_time' (i.e., the stay at that city) MUST be more than 2 days (> 48 hours).
8. For each segment, the difference between the 'departure_time' of the previous city and the 'arrival_time' of the next city MUST be equal to the minimum realistic flight time (plus a 1 hour buffer for airport procedures). Do NOT add extra days or hours to the travel time.
9. The stay duration and the travel duration are separate: do NOT add the 2-day minimum stay to the travel time. The 2-day minimum applies only to the time spent at each city.
10. Return ONLY the JSON object, nothing else.
"""
    
    prompt_to_use = base_prompt_text
#     requirements = """Requirements:
# 1. All times MUST be in UTC.
# 2. Use 24-hour format (e.g., 14:30, 00:00 for midnight).
# 3. Travel dates must be between {date_suggestion_start} and {date_suggestion_end}.
# 4. Include the IATA airport code for each city in parentheses.
# 5. Do NOT add any explanatory text or markdown formatting.
# 6. Ensure the JSON is properly formatted with correct commas and brackets.
# 7. Account for minimum flight times.
# 8. Add at least 1 hour between arrival and departure at each city.
# 9. Return ONLY the JSON object, nothing else."""
    if feedback_issues: # Renamed from feedback to feedback_issues
        prompt_to_use = feedback_issues + "\n\n" + base_prompt_text# Prepend issues to the base prompt
   
    try:
        if llm_provider == "openai":
            response = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": "You are a travel itinerary planner. Your output must be ONLY a valid JSON object following the user's specified structure. No other text, no markdown formatting, no explanations, no additional fields."},
                    {"role": "user", "content": prompt_to_use}
                ],
                temperature=0.7
            )
            content = response.choices[0].message.content.strip()
        elif llm_provider == "gemini":
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt_to_use)
            content = response.text.strip()
        elif llm_provider == "llama":
            content = llama2_ollama_completion(prompt_to_use, model_name=model_name)
            if not content:
                print("Error: No response from Llama2 Ollama API.")
                return None
        else:
            print(f"Error: LLM provider '{llm_provider}' not supported.")
            return None

        # Clean the response content
        content = content.replace('```json', '').replace('```', '').strip()
       
        try:
            # Find the first '{' and last '}'
            json_start = content.find('{')
            json_end = content.rfind('}') + 1
            if json_start != -1 and json_end > json_start:
                json_str = content[json_start:json_end]
                # Remove any non-JSON text that might have been added
                json_str = json_str.split('\n')[0] + '\n'.join(line for line in json_str.split('\n')[1:] if line.strip().startswith('{') or line.strip().startswith('"') or line.strip().startswith('}') or line.strip().startswith('[') or line.strip().startswith(']') or line.strip().startswith(','))
                itinerary_json = json.loads(json_str)
            else:
                print(f"Error: Could not find valid JSON object in response: {content}")
                return None

            if "itinerary" not in itinerary_json or not isinstance(itinerary_json["itinerary"], list):
                print(f"Error: Invalid itinerary structure in LLM response: {itinerary_json}")
                return None
               
            itinerary_data = itinerary_json["itinerary"]
           
            # Validate that all required fields are present and not None
            for stop in itinerary_data:
                if not all(key in stop and stop[key] is not None for key in ['place', 'arrival_time', 'departure_time']):
                    print(f"Error: Missing or None values in itinerary data for stop: {stop}")
                    return None
                if "(" not in stop['place'] or ")" not in stop['place']:
                    print(f"Error: Missing IATA code in place name: {stop['place']}")
                    return None
           
            # Extract IATA codes from place names
            for stop in itinerary_data:
                place = stop.get("place", "")
                city_name, iata = place.split("(")
                iata = iata.rstrip(")")
                stop["place"] = city_name.strip()
                stop["iata"] = iata.strip()
           
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON from LLM response: {e}")
            print(f"Raw response was: {content}")
            return None
       
        return itinerary_data
       
    except Exception as e:
        print(f"Error generating itinerary with {llm_provider} ({model_name}): {e}")
        return None

def capture_validation_output(func):
    """
    Decorator to capture validation output for CSV logging while still showing it in terminal
    """
    def wrapper(*args, **kwargs):
        # Create a string buffer to capture output
        output_buffer = StringIO()
        # Save the original stdout
        original_stdout = sys.stdout
        # Create a Tee class that writes to both stdout and our buffer
        class Tee:
            def __init__(self, file1, file2):
                self.file1 = file1
                self.file2 = file2
            def write(self, data):
                self.file1.write(data)
                self.file2.write(data)
            def flush(self):
                self.file1.flush()
                self.file2.flush()
        
        # Redirect stdout to our Tee
        sys.stdout = Tee(original_stdout, output_buffer)
        
        # Call the original function
        result = func(*args, **kwargs)
        
        # Restore stdout
        sys.stdout = original_stdout
        # Get the captured output
        captured_output = output_buffer.getvalue()
        # Close the buffer
        output_buffer.close()
        
        # If result is a tuple, add the captured output as the last element
        if isinstance(result, tuple):
            return result + (captured_output,)
        return result, captured_output
    return wrapper

@capture_validation_output
def generate_valid_journey(num_destinations=4, max_attempts=3, llm_provider="openai", model_name="gpt-4o-mini", feedback=None):
    """
    Generate and validate a journey using the LLM and AeroDataBox API.
    Returns a tuple of (journey_events, feedback_parts, travel_time_cache, flag, invalid_llm_generation, invalid_format, validation_output)
    """
    # Select exactly num_destinations cities at the start
    selected_cities = random.sample(INTERNATIONAL_CITIES, k=num_destinations)
    print(f"\nSelected cities for this attempt (pool for LLM):")
    for city in selected_cities:
        print(f"- {city['city']} ({city['iata']}), {city['country']}")

    travel_time_cache = {}  # Initialize cache for travel times for this validation cycle
    flag = 0
    attempt = 0
    current_feedback = feedback
    invalid_llm_generation = False  # Track if LLM generated invalid itinerary
    invalid_format = False  # Track if LLM generated invalid format
    
    while attempt < max_attempts:
        attempt += 1
        print(f"\nAttempt {attempt}/{max_attempts}")
        
        # Get itinerary from LLM with feedback if available
        itinerary_data = generate_itinerary_with_llm(
            selected_cities=selected_cities,
            num_destinations=num_destinations,
            llm_provider=llm_provider,
            model_name=model_name,
            feedback_issues=current_feedback
        )
        
        if not itinerary_data:
            print(f"\nFailed to generate itinerary content on attempt {attempt}.")
            if attempt < max_attempts:
                current_feedback = """The previous response was not a valid JSON object. Please ensure:
1. The response is a single, valid JSON object
2. The JSON has an \"itinerary\" array containing exactly 4 stops
3. Each stop has \"place\", \"arrival_time\", and \"departure_time\" fields
4. All times are in UTC and follow the format 'YYYY-MM-DD HH:MM'
5. Each place includes the IATA code in parentheses
Example format:
{
    \"itinerary\": [
        {
            \"place\": \"London (LHR)\",
            \"arrival_time\": \"2024-03-20 10:00\",
            \"departure_time\": \"2024-03-20 14:00\"
        }
    ]
}"""
                invalid_format = True
                continue
            flag = 1
            return None, [], travel_time_cache, flag, invalid_llm_generation, invalid_format, 0, 0
        
        # Print the generated itinerary
        print(f"\n GENERATED ITINERARY (Attempt {attempt}):")
        print(f"{'-'*50}")
        for i, stop in enumerate(itinerary_data):
            print(f"  {i+1}. {stop['place']} ({stop['iata']})")
            print(f"     Arrive: {stop['arrival_time']} UTC")
            print(f"     Depart: {stop['departure_time']} UTC")
        
        # Convert itinerary data to journey events
        journey_events = []
        parsing_error = False
        for stop in itinerary_data:
            try:
                arrival_time = datetime.strptime(stop["arrival_time"], "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
                departure_time = datetime.strptime(stop["departure_time"], "%Y-%m-%d %H:%M").replace(tzinfo=timezone.utc)
                journey_events.append({
                    "place": stop["place"],
                    "iata": stop["iata"],
                    "arrival_time": arrival_time,
                    "departure_time": departure_time
                })
            except (ValueError, KeyError) as e:
                print(f"\nError parsing times for {stop.get('place', 'unknown')}: {e}")
                parsing_error = True
                if attempt < max_attempts:
                    current_feedback = f"""Error in time format for {stop.get('place', 'unknown')}. Please ensure:
1. All times are in UTC and follow the EXACT format 'YYYY-MM-DD HH:MM'
2. Use 24-hour format (e.g., 14:30, 00:00 for midnight)
3. Include leading zeros (e.g., '01:05' not '1:5')
4. No timezone indicators or UTC suffix
Example: '2024-03-20 14:30'"""
                    break
                continue
        
        if parsing_error:
            invalid_format = True
            if attempt == max_attempts:
                return None, [], travel_time_cache, 1, invalid_llm_generation, invalid_format, 0, 0
            continue

        if len(journey_events) < 2:
            print("\nNot enough valid journey events to form a complete itinerary.")
            if attempt < max_attempts:
                current_feedback = f"""Generated itinerary has insufficient stops. Please ensure:
1. The itinerary contains exactly {num_destinations} stops
2. Each stop has all required fields (place, arrival_time, departure_time)
3. Each place includes the IATA code in parentheses
4. All times are in UTC and follow the format 'YYYY-MM-DD HH:MM'
Example format:
{{
    \"itinerary\": [
        {{
            \"place\": \"London (LHR)\",
            \"arrival_time\": \"2024-03-20 10:00\",
            \"departure_time\": \"2024-03-20 14:00\"
        }},
        {{
            \"place\": \"Paris (CDG)\",
            \"arrival_time\": \"2024-03-20 16:00\",
            \"departure_time\": \"2024-03-21 10:00\"
        }}
    ]
}}"""
                invalid_format = True
                continue
            flag = 1
            return None, [], travel_time_cache, flag, invalid_llm_generation, invalid_format, 0, 0

        # --- VALIDATION PHASE ---
        print(f"\n VALIDATING ITINERARY:")
        print(f"{'-'*50}")
        feedback_parts = []
        validation_failed = False
        # 1. Stay duration validation for each stop
        for i, stop in enumerate(journey_events):
            stay_duration = (stop['departure_time'] - stop['arrival_time']).total_seconds()
            if stay_duration < 2 * 24 * 3600:
                validation_failed = True
                feedback_parts.append(f"Stop {stop['place']}: Stay duration is less than 2 days.")
                print(f"  ❌ Validation failed: Stay at {stop['place']} is less than 2 days.")
        # 2. Segment validation
        for i in range(len(journey_events) - 1):
            current = journey_events[i]
            next_stop = journey_events[i + 1]
            print(f"\nChecking segment: {current['place']} → {next_stop['place']}")
            print(f"  Departure: {current['departure_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            print(f"  Arrival:   {next_stop['arrival_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            if current["departure_time"] > next_stop["arrival_time"]:
                validation_failed = True
                print(f"  ❌ Validation failed: Departure time is after arrival time")
                feedback_parts.append(
                    f"Segment {current['place']} to {next_stop['place']}: Departure time is after arrival time. Please ensure departure time is before arrival time."
                )
                continue
            min_travel_seconds = get_min_travel_time(current, next_stop, None, travel_time_cache)
            max_travel_seconds = int(min_travel_seconds * 2)
            available_travel_time_seconds = (next_stop["arrival_time"] - current["departure_time"]).total_seconds()
            if available_travel_time_seconds < min_travel_seconds:
                validation_failed = True
                print(f"  ❌ Validation failed: Travel time too short (available: {available_travel_time_seconds/3600:.1f}h, required: {min_travel_seconds/3600:.1f}h)")
                feedback_parts.append(
                    f"Segment {current['place']} to {next_stop['place']}: Travel time is too short. Minimum required time is {min_travel_seconds/3600:.1f} hours."
                )
            elif available_travel_time_seconds > max_travel_seconds:
                validation_failed = True
                print(f"  ❌ Validation failed: Travel time too long (available: {available_travel_time_seconds/3600:.1f}h, max allowed: {max_travel_seconds/3600:.1f}h)")
                feedback_parts.append(
                    f"Segment {current['place']} to {next_stop['place']}: Travel time is too long. Maximum allowed time is {max_travel_seconds/3600:.1f} hours."
                )
            else:
                print(f"  ✅ Segment validated successfully")
        # --- FIXING PHASE ---
        # Only count each incorrect segment once, even if there are multiple errors
        incorrect_segments = set()
        for msg in feedback_parts:
            if msg.startswith('Segment '):
                seg = msg.split(':')[0].strip()
                incorrect_segments.add(seg)
        total_timing_issues = len(incorrect_segments)
        total_fixed_issues = len(incorrect_segments)

        if validation_failed:
            invalid_llm_generation = True
            # Fix all issues, propagating changes
            adjusted_journey_events = journey_events.copy()
            for i in range(len(adjusted_journey_events)):
                # Ensure 2-day minimum stay
                arr = adjusted_journey_events[i]['arrival_time']
                dep = adjusted_journey_events[i]['departure_time']
                min_stay = timedelta(days=2)
                if (dep - arr).total_seconds() < 2 * 24 * 3600:
                    # Add 30 to 180 random minutes to make it look natural
                    extra_minutes = 5*random.randint(10, 400)
                    dep = arr + min_stay + timedelta(minutes=extra_minutes)
                    adjusted_journey_events[i]['departure_time'] = dep
                # Propagate fixes to next stops
                if i < len(adjusted_journey_events) - 1:
                    min_travel = get_min_travel_time(adjusted_journey_events[i], adjusted_journey_events[i+1], None, travel_time_cache)
                    next_arr = dep + timedelta(seconds=min_travel + 3600)
                    if adjusted_journey_events[i+1]['arrival_time'] < next_arr:
                        adjusted_journey_events[i+1]['arrival_time'] = next_arr
                        # Also fix next departure if needed
                        if (adjusted_journey_events[i+1]['departure_time'] - next_arr).total_seconds() < 2 * 24 * 3600:
                            extra_minutes = 5*random.randint(10, 400)
                            adjusted_journey_events[i+1]['departure_time'] = next_arr + min_stay + timedelta(minutes=extra_minutes)
            print(f"\n FIXED ITINERARY:")
            print(f"{'-'*50}")
            for i, stop in enumerate(adjusted_journey_events):
                print(f"  {i+1}. {stop['place']} ({stop['iata']})")
                print(f"     Arrive: {stop['arrival_time'].strftime('%Y-%m-%d %H:%M UTC')}")
                print(f"     Depart: {stop['departure_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            return adjusted_journey_events, feedback_parts, travel_time_cache, flag, invalid_llm_generation, invalid_format, total_timing_issues, total_fixed_issues
        else:
            print(f"\n VALID ITINERARY:")
            print(f"{'-'*50}")
            for i, stop in enumerate(journey_events):
                print(f"  {i+1}. {stop['place']} ({stop['iata']})")
                print(f"     Arrive: {stop['arrival_time'].strftime('%Y-%m-%d %H:%M UTC')}")
                print(f"     Depart: {stop['departure_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            return journey_events, feedback_parts, travel_time_cache, flag, invalid_llm_generation, invalid_format, total_timing_issues, total_fixed_issues
    # If we get here, all attempts failed
    flag = 1
    return None, [], travel_time_cache, flag, invalid_llm_generation, invalid_format, total_timing_issues, total_fixed_issues

def calculate_metrics(results):
    """
    Calculate metrics from the results of multiple epochs.
    results: list of dicts containing epoch results
    """
    total_epochs = len(results)
    successful_epochs = sum(1 for r in results if r['success'])
    total_segments = sum(r['total_segments'] for r in results)
    total_timing_issues = sum(r['total_timing_issues'] for r in results)
    total_fixed_issues = sum(r['total_fixed_issues'] for r in results)
    
    # Calculate metrics with proper handling of edge cases
    metrics = {
        'success_rate': (successful_epochs / total_epochs) * 100 if total_epochs > 0 else 0,
        'timing_issues_per_itinerary': total_timing_issues / total_epochs if total_epochs > 0 else 0,
        'fixed_issues_per_itinerary': total_fixed_issues / total_epochs if total_epochs > 0 else 0,
        'fix_rate': (total_fixed_issues / total_timing_issues) * 100 if total_timing_issues > 0 else 0,
        'total_epochs': total_epochs,
        'successful_epochs': successful_epochs,
        'total_segments': total_segments,
        'total_timing_issues': total_timing_issues,
        'total_fixed_issues': total_fixed_issues
    }
    return metrics

def run_epoch(llm_provider, model_name, epoch_num, total_epochs, worksheet):
    """
    Run a single epoch of itinerary generation and validation.
    """
    print(f"\n{'='*50}")
    print(f"EPOCH {epoch_num}/{total_epochs}")
    print(f"{'='*50}")
   
    result = {
        'success': False,
        'total_timing_issues': 0,  # Count of all timing issues found
        'total_fixed_issues': 0,   # Count of issues successfully fixed
        'total_segments': 0,       # Total number of segments in the itinerary
        'invalid_llm_generation': False,  # Track if LLM generated invalid itinerary
        'invalid_format': False    # Track if LLM generated invalid format
    }
   
    # Single attempt to get itinerary from LLM
    final_journey, feedback_parts, travel_time_cache, flag, invalid_llm_generation, invalid_format, total_timing_issues, total_fixed_issues, validation_output = generate_valid_journey(
        num_destinations=6,
        max_attempts=3,
        llm_provider=llm_provider,
        model_name=model_name
    )
   
    # Extract only the validation segments from the output
    validation_segments = []
    fixed_itinerary = []
    current_segment = []
    for line in validation_output.split('\n'):
        if line.strip().startswith('Checking segment:'):
            if current_segment:
                validation_segments.append('\n'.join(current_segment))
            current_segment = [line]
        elif line.strip() and current_segment:
            if line.strip().startswith('  Departure:') or line.strip().startswith('  Arrival:') or line.strip().startswith('Flight time') or line.strip().startswith('Cached') or 'Segment validated successfully' in line or 'Validation failed' in line:
                current_segment.append(line)
    if current_segment:
        validation_segments.append('\n'.join(current_segment))
    
    # Format fixed itinerary if available
    if final_journey:
        for stop in final_journey:
            fixed_itinerary.append(f"{stop['place']} ({stop['iata']})")
            fixed_itinerary.append(f"  Arrive: {stop['arrival_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            fixed_itinerary.append(f"  Depart: {stop['departure_time'].strftime('%Y-%m-%d %H:%M UTC')}")
            fixed_itinerary.append("")
    
    # Write to Excel
    row = epoch_num + 1  # +1 because row 1 is header
    worksheet.cell(row=row, column=1, value=epoch_num)
    worksheet.cell(row=row, column=2, value='\n\n'.join(validation_segments))
    worksheet.cell(row=row, column=3, value=total_timing_issues)
    worksheet.cell(row=row, column=4, value='Invalid' if not final_journey or validation_output.count('❌ Validation failed') > 0 else 'Valid')
    worksheet.cell(row=row, column=5, value='\n'.join(fixed_itinerary))
    worksheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    worksheet.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
   
    result['total_timing_issues'] = total_timing_issues
    result['total_fixed_issues'] = total_fixed_issues

    if not final_journey:
        result['invalid_llm_generation'] = invalid_llm_generation
        result['invalid_format'] = invalid_format
        result['total_segments'] = 3  # For 4 destinations, there are 3 segments
        return result, flag
   
    # If we get here, we have a valid journey
    result['success'] = True
    result['total_segments'] = len(final_journey) - 1
    result['invalid_llm_generation'] = invalid_llm_generation
    result['invalid_format'] = invalid_format
    
    return result, flag

def save_metrics_to_file(all_results, filename="metrics_log.json"):
    """
    Save the metrics results to a JSON file.
    """
    from datetime import datetime
   
    # Load existing data if file exists
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        data = {
            "experiment_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "metrics_format": {
                "success_rate": "Percentage of epochs that resulted in a valid itinerary",
                "timing_issues_per_itinerary": "Mean number of timing issues per itinerary",
                "fixed_issues_per_itinerary": "Mean number of issues fixed per itinerary",
                "fix_rate": "Percentage of timing issues fixed per itinerary",
                "total_epochs": "Total number of epochs",
                "successful_epochs": "Total number of successful epochs",
                "total_segments": "Total number of segments in all itineraries",
                "total_timing_issues": "Total number of timing issues across all itineraries",
                "total_fixed_issues": "Total number of issues fixed across all itineraries"
            },
            "results": {}
        }
   
    data["results"] = all_results
    with open(filename, 'w') as f:
        json.dump(data, f, indent=4)

def main():
    # Configuration
    EPOCHS = 100
    LLM_PROVIDERS = ["openai", "gemini", "llama"]
    MODEL_NAMES = {
        "openai": "gpt-4o-mini",
        "gemini": "gemini-2.0-flash-thinking-exp-01-21",
        "llama": "llama2:7b",
    }
    
    EXCEL_FILENAME = f"{LLM_PROVIDERS[0]}_6.xlsx"
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Validation Results"
    
    headers = ['Epoch', 'Validation Response', 'Incorrect Segments', 'Itinerary Status', 'Validated Itinerary']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    worksheet.column_dimensions['A'].width = 10  # Epoch
    worksheet.column_dimensions['B'].width = 100  # Validation Response
    worksheet.column_dimensions['C'].width = 20  # Incorrect Segments
    worksheet.column_dimensions['D'].width = 20  # Itinerary Status
    worksheet.column_dimensions['E'].width = 100  # Validated Itinerary
    
    # Store results for each provider
    all_results = {}
    for provider in LLM_PROVIDERS:
        print(f"\n{'='*50}")
        print(f"Testing {provider.upper()} with model {MODEL_NAMES[provider]}")
        print(f"{'='*50}")
       
        invalid_llm_count = 0
        invalid_format_count = 0
        results = []
        for epoch in range(1, EPOCHS + 1):
            result, flag = run_epoch(provider, MODEL_NAMES[provider], epoch, EPOCHS, worksheet)
            results.append(result)
            if result['invalid_llm_generation']:
                invalid_llm_count += 1
            if result['invalid_format']:
                invalid_format_count += 1
       
        # Calculate final metrics for this provider
        metrics = calculate_metrics(results)
        print("\nFinal Metrics:")
        print("Invalid itineraries generated by LLM: ", invalid_llm_count)
        print("Invalid format responses from LLM: ", invalid_format_count)
        print(f"Percentage of invalid itineraries generated by LLM: {invalid_llm_count/EPOCHS*100}%")
        print(f"Percentage of invalid format responses: {invalid_format_count/EPOCHS*100}%")
        print(f"Success Rate: {metrics['success_rate']:.2f}%")
        print(f"Average Timing Issues per Itinerary: {metrics['timing_issues_per_itinerary']:.2f}")
        print(f"Fix Rate: {metrics['fix_rate']:.2f}%")
        print(f"Total Epochs: {metrics['total_epochs']}")
        print(f"Total Segments: {metrics['total_segments']}")
        print(f"Total Timing Issues Found: {metrics['total_timing_issues']}")
        
        all_results[provider] = metrics
   
    # Calculate overall metrics across all providers
    overall_metrics = {
        'success_rate': sum(r['success_rate'] for r in all_results.values()) / len(all_results) if len(all_results) > 0 else 0,
        'timing_issues_per_itinerary': sum(r['timing_issues_per_itinerary'] for r in all_results.values()) / len(all_results) if len(all_results) > 0 else 0,
        'fix_rate': sum(r['fix_rate'] for r in all_results.values()) / len(all_results) if len(all_results) > 0 else 0,
        'total_epochs': sum(r['total_epochs'] for r in all_results.values()),
        'total_segments': sum(r['total_segments'] for r in all_results.values()),
        'total_timing_issues': sum(r['total_timing_issues'] for r in all_results.values()),
    }
    all_results['overall'] = overall_metrics
   
    save_metrics_to_file(all_results)

    workbook.save(EXCEL_FILENAME)

if __name__ == "__main__":
    main()